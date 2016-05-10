import json
import time

import datetime
from collections import defaultdict
import random

from traveling_salesman.TSPIndividual import TSPIndividual


class EvolutionWorld:
    def __init__(self, parent_population_size=100, tournament_e=0.1, mutation_rate=0.8, crossover_rate=0.5):
        # Set initial variables
        self.mutation_rate = mutation_rate
        self.crossover_rate = crossover_rate
        self.tournament_e = tournament_e
        self.resulting_population, self.parent_population, self.offspring_population = set(), set(), set()
        self.parent_population_size = parent_population_size
        self.distance, self.rank = dict(), dict()
        self.front_list = list()
        self.generations_run = 0

        # Load travel costs from disk
        self.cost_dict = dict()
        self.cost_dict["cost"], self.cost_dict["distance"] = self.load_travel_data()

        # Generate an initial population and put it in the offspring pool
        self.generate_initial_population()

    def generate_initial_population(self):
        """
        Generates an initial population twice the size of the parent pool.
        """
        self.offspring_population = {TSPIndividual(cost_dict=self.cost_dict) for i in
                                     range(self.parent_population_size * 2)}

    def main_loop(self):
        """
        Runs one complete round of evolution.
        """
        self.assess_offspring_fitness()
        self.resulting_population = self.parent_population.union(self.offspring_population)
        new_parents = set()
        distance = dict()

        self.front_list, rank = self.fast_non_dominated_sort(self.resulting_population)

        for i in range(len(self.front_list)):
            front_i = self.front_list[i]  # Get current front
            distance.update(self.crowding_distance_assignment(front_i))  # Update their distances

            if len(new_parents) + len(front_i) > self.parent_population_size:
                # Not enough room for all members of front
                sorted_by_distance = self.crowded_comparison_sort(front_i, rank, distance)
                stop_index = self.parent_population_size - len(new_parents)  # First index to not be included
                new_parents = new_parents.union(sorted_by_distance[:stop_index])  # Add to parent pool
                break

            new_parents = new_parents.union(front_i)

        self.parent_population = new_parents
        self.distance = distance
        self.rank = rank

        self.offspring_population = self.make_new_offspring(self.parent_population)
        self.generations_run += 1

    def make_new_offspring(self, parent_population):
        """
        Fills the offspring pool with new offspring.
        :param parent_population: The parent population used to generate offspring.
        :return:
        """
        new_offspring = set()
        parent_population_list = list(parent_population)

        while len(new_offspring) < self.parent_population_size:
            change_happened = False
            first_individual = random.choice(parent_population_list)
            second_individual = random.choice(parent_population_list)
            chosen_individual = self.binary_tournament_select(first_individual, second_individual)

            if random.random() < self.crossover_rate:
                change_happened = True
                first_individual = random.choice(parent_population_list)
                second_individual = random.choice(parent_population_list)
                chosen_mate = self.binary_tournament_select(first_individual, second_individual)
                child_genotype = chosen_individual.mate_with(chosen_mate)
                chosen_individual = TSPIndividual(cost_dict=self.cost_dict, genotype=child_genotype)

            if random.random() < self.mutation_rate:
                change_happened = True
                mutated_genotype = chosen_individual.get_mutated_genotype()
                chosen_individual = TSPIndividual(cost_dict=self.cost_dict, genotype=mutated_genotype)

            if change_happened:
                new_offspring.add(chosen_individual)

        return new_offspring

    def binary_tournament_select(self, one, other):
        """
        Holds a binary tournament between two parents contending to make offspring.
        :param one: An individual
        :param other: Another individual
        :return: The winner of the tournament
        """
        ranking = self.crowded_comparison_sort([one, other], self.rank, self.distance)

        if random.random() < self.tournament_e:
            return ranking[1]
        return ranking[0]

    def assess_offspring_fitness(self):
        for individual in self.offspring_population:
            individual.assess_fitness()

    def run_for_x_generations(self, generations=1000):
        for i in range(generations):
            if i % 100 == 0:
                print(self.generations_run)
            self.main_loop()

    def log_fronts(self):
        """
        Logs all the fronts to a json file.
        """
        current_time = time.time()
        datestring = datetime.datetime.fromtimestamp(current_time).strftime('%Y-%m-%d-%H:%M:%S')

        log_list = []
        for front in self.front_list:
            this_front_fitnesses = []
            for individual in front:
                this_front_fitnesses.append(individual.fitnesses)

            key1, key2 = tuple(individual.fitnesses.keys())

            this_front_fitnesses.sort(key=lambda x: (x[key1], -x[key2]))

            log_list.append(this_front_fitnesses)

        log_name = datestring + '-P' + str(self.parent_population_size) + '-G' + str(self.generations_run) + '-M' + str(
            self.mutation_rate) + '-C' + str(self.mutation_rate) + '-E' + str(self.tournament_e) + '.log'

        log_path = 'traveling_salesman/logs/'

        with open(log_path + log_name, 'w') as log_file:
            json.dump(log_list, log_file, indent=2)

    @staticmethod
    def crowded_comparison_sort(population, rank, distance):
        """
        Sorts a list by ascending rank and decreasing distance.

        :param population: A population set.
        :param rank: A dictionary of ranks.
        :param distance: A distance
        :return:
        """
        population_as_list = list(population)
        population_as_list.sort(key=lambda x: (rank[x], -distance[x]))
        return population_as_list

    @staticmethod
    def fast_non_dominated_sort(population):
        """
        Sorts the population into non-dominated fronts.

        :param population: A population of individuals that have the dominates property.
        :return:
        """
        domination_set = defaultdict(set)
        domination_count = defaultdict(int)
        rank = dict()
        f = [set()]

        for p in population:
            for q in population:
                if q == p:
                    continue
                if p.dominates(q):
                    domination_set[p].add(q)  # Add q to the set of solutions dominated by p
                elif q.dominates(p):
                    domination_count[p] += 1  # Increment the domination counter of p

            if domination_count[p] == 0:  # p belongs to the first front
                rank[p] = 0  # EDIT: Ranks start at 0 instead of 1
                f[0].add(p)

        i = 0  # Initialize the front counter. EDIT: Starts at 0

        while f[i]:  # Set contains elements
            next_front = set()  # Replacement for Q in pseudo code
            for p in f[i]:
                for q in domination_set[p]:
                    domination_count[q] -= 1

                    if domination_count[q] == 0:  # q belongs to next front
                        rank[q] = i + 1
                        next_front.add(q)

            i += 1
            f.append(next_front)

        f.pop(-1)  # Last set is always empty, so remove it

        return f, rank

    @staticmethod
    def crowding_distance_assignment(population):
        """
        Assigns a crowding distance to a population in the same front.

        :param population: A population of individuals.
        :return: A dict of the crowding distances.
        """
        distance = defaultdict(int)

        population_list = list(population)
        population_size = len(population_list)

        objectives = population_list[0].fitnesses.keys()

        for m in objectives:
            population_list.sort(key=lambda x: x.fitnesses[m])
            min_element = population_list[0]
            max_element = population_list[population_size - 1]

            f_min = min_element.fitnesses[m]
            f_max = max_element.fitnesses[m]

            span = f_max - f_min

            if span == 0:  # The population does not vary in this property
                continue

            distance[min_element] = distance[max_element] = float("inf")

            for i in range(1, population_size - 1):
                previous_element = population_list[i - 1]
                current_element = population_list[i]
                next_element = population_list[i + 1]

                current_distance = (next_element.fitnesses[m] - previous_element.fitnesses[m]) / span

                distance[current_element] += current_distance

        return distance

    @classmethod
    def load_travel_data(cls):
        import pickle
        try:
            with open("traveling_salesman/travel_data/Distance.pickle", "rb") as d_file:
                distances = pickle.load(d_file)
            with open("traveling_salesman/travel_data/Cost.pickle", "rb") as c_file:
                costs = pickle.load(c_file)
        except IOError:
            from openpyxl import load_workbook

            costs = load_workbook('traveling_salesman/travel_data/Cost.xlsx', read_only=True)
            distances = load_workbook('traveling_salesman/travel_data/Distance.xlsx', read_only=True)

            costs = cls.convert_sheet_to_map(costs["Sheet1"])
            distances = cls.convert_sheet_to_map(distances["Sheet1"])

            with open("traveling_salesman/travel_data/Distance.pickle", "wb") as d_file:
                pickle.dump(distances, d_file)

            with open("traveling_salesman/travel_data/Cost.pickle", "wb") as c_file:
                pickle.dump(costs, c_file)

        return costs, distances

    @staticmethod
    def convert_sheet_to_map(sheet):
        costs = dict()
        for row in range(2, sheet.max_row + 1):
            for column in range(2, row + 1):
                data = sheet.cell(row=row, column=column).value

                actual_first_city_number = column - 1
                actual_second_city_number = row - 1
                first_way = (actual_first_city_number, actual_second_city_number)
                second_way = (actual_second_city_number, actual_first_city_number)

                costs[first_way] = data
                costs[second_way] = data

        return costs
